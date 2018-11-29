Author： RichardHuang
Time: 2018/11/25--
Aim: 爬取拉勾的最新招聘信息，具体是上海-测试-150人以上公司-3年以下工作经验，同时保存至表格。保持登陆cookies，在HR端的“谁看过我”留下记录，前排混眼熟
Functions:  login_in(phonenum, password)  登陆函数，保存cookies
            search(keyword) 进入拉勾主页后搜索关键字及选项
            get_job_url() 解析搜索的网页，获取每个工作的url
            get_job_info(url) 打开url，爬取工作要求及岗位职责
            save_to_excel(message) 保存工作信息至表格
Next Plan： 加入高德地图API，在地图上标注公司位置，直观显示距离
            清洗数据，统计薪资与行业、公司规模是否存在某种关系
"""

# -*- coding: UTF-8 -*-

import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import Workbook

driver = webdriver.Chrome()
driver.get('https://www.lagou.com/')
cookies = driver.get_cookies()


def login_in(phonenum, password):
    driver.get('https://www.lagou.com/')
    assert "拉勾网" in driver.title
    ActionChains(driver).move_to_element(city).click().perform()

    #  登录方法
    login_link = WebDriverWait(driver,5).until(EC.visibility_of_element_located((By.LINK_TEXT, '登录')))
    ActionChains(driver).move_to_element(login_link).click().perform()

    driver.find_element_by_xpath(u"(.//*[normalize-space(text()) and normalize-space(.)='验证码登录'])[1]/following::input[1]").send_keys(phonenum)
    driver.find_element_by_xpath(u"(.//*[normalize-space(text()) and normalize-space(.)='验证码登录'])[1]/following::input[2]").send_keys(password)
    driver.find_element_by_xpath(u"(.//*[normalize-space(text()) and normalize-space(.)='忘记密码？'])[1]/following::input[1]").click()

    time.sleep(7)  # 暂时手动确认验证码
    driver.refresh()  # 登录后刷新获取cookies
    assert "拉勾网" in driver.title
    cookies = driver.get_cookies()
    # print(cookies)


def search(keyword):
    search_text = WebDriverWait(driver,8).until(EC.visibility_of_element_located((By.ID, 'search_input')))
    search_text.send_keys(keyword)
    driver.find_element_by_id('search_button').click()
    driver.find_element_by_link_text(experience_1).click()
    time.sleep(0.6)
    driver.find_element_by_link_text(experience_2).click()
    time.sleep(0.6)
    driver.find_element_by_link_text(recent).click()
    time.sleep(0.6)
    driver.find_element_by_link_text(com_size_1).click()
    time.sleep(0.6)
    driver.find_element_by_link_text(com_size_2).click()
    time.sleep(0.6)
    driver.find_element_by_link_text(com_size_3).click()


def get_job_url():
    time.sleep(1)
    soup_1 = BeautifulSoup(driver.page_source, "lxml")
    page_num = soup_1.find('span', {'class': 'span totalNum'}).string.strip()

    for i in range(int(1)):  # 所有页数的工作信息
        soup = BeautifulSoup(driver.page_source, "html.parser")
        com_list = soup.find('div', {'id': 's_position_list'})
        # print(driver.page_source)

        for job_info in com_list.findAll('li'):
            # print(job_info)
            job_url = job_info.find('a', {'class': 'position_link'})
            if job_url is not None:  # 网页中错误的标签抛弃
                get_job_info(job_url['href'])
            else:
                continue
        driver.find_element_by_class_name('pager_next ').click()
        time.sleep(1)


def get_job_info(url):
    detail = webdriver.Chrome()
    detail.get(url)
    for cookie in cookies:
        detail.add_cookie(cookie)
    time.sleep(1)
    detail.refresh()
    job_messages = []  # 用来保存爬取的工作信息
    if ('互联网招聘求职网' in detail.title) or ('www.lagou.com/' in detail.title):
        get_job_info(url)
    else:
        job_desc = BeautifulSoup(detail.page_source, "html.parser")  # 解析工作详情网页
        job_com = job_desc.find('div',{'class': 'company'}).string.strip()
        job_name = job_desc.find('span',{'class': 'name'}).string.strip()
        job_all_req = job_desc.find('dd',{'class': 'job_request'}).findAll('span')  # 获取薪资、地址、经验要求

        job_messages.append(job_name)
        job_messages.append(job_com)
        for job_req in job_all_req:
            job_messages.append(job_req.contents[0])
        # print(job_messages)

        job_bt = job_desc.find('dd',{'class': 'job_bt'}).findAll('p') # 获取职位描述
        job_address = job_desc.find('div',{'class': 'work_addr'}) # 获取工作地址
        #print(job_address.contents[6])
        job_messages.append(job_bt)
        save_to_excel(job_messages)
        print(job_messages)
    detail.quit()


wb = Workbook(write_only=True)
ws = []
ws.append(wb.create_sheet(title='测试'))
ws[0].append(['工作名称', '公司', '薪资', '经验要求', '学历要求', '全职与否', '岗位职责'])


def save_to_excel(message):
    ws[0].append([message[0], message[1], message[2], message[4], message[5], message[6], str(message[7])])


if __name__ == '__main__':
    where = '上海站'  # login_in 内部会选择求职地点
    login_in('usename', 'password')  # 自己的账号和密码
    job_want = '测试'
    recent = '最新'
    experience_1 = '3年及以下'
    experience_2 = '不要求'
    com_size_1 = '150-500人'
    com_size_2 = '500-2000人'
    com_size_3 = '2000人以上'  # 以上参数用于search() 工作名称及相关要求
    search(job_want)
    get_job_url()
    wb.save('job_list.xlsx')
    driver.quit()
    #os.system("pause");